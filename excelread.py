import openpyxl
from openpyxl.utils import column_index_from_string
from robot.api.deco import keyword

@keyword('Read Excel File')
def read_excel_file(file_path, sheet_name, row_number):
    if row_number < 3:
        row_number = 3

    workbook = openpyxl.load_workbook(file_path)

    try:
        sheet = workbook[sheet_name]
    except:
        raise Exception(f"sheet {sheet_name} not found")

    header = [cell.value.lower() if cell.value is not None else "" for cell in sheet[2]]
    row = [cell.value if cell.value is not None else "" for cell in sheet[row_number]]
    return dict(zip(header, row))

@keyword('Read All Excel Content')
def read_all_excel_content(file_path, sheet_name, header_line = 2):
    workbook = openpyxl.load_workbook(file_path)
    try:
        sheet = workbook[sheet_name]
    except:
        raise Exception(f"sheet {sheet_name} not found")
    header = [str(cell.value).lower() if cell.value is not None else "" for cell in sheet[header_line] if cell.value]
    all_rows = []
    for row in sheet.iter_rows(min_row=header_line + 1, values_only=True):
        row_dict = dict(zip(header, row))
        all_rows.append(row_dict)
    return all_rows

@keyword('Write Column If Condition Matches')
def write_column_if_condition_matches(file_path, sheet_name, condition_col, condition_val, target_col, target_val, header_line=2):
    workbook = openpyxl.load_workbook(file_path)
    try:
        sheet = workbook[sheet_name]
    except:
        raise Exception(f"sheet {sheet_name} not found")
    header = [str(cell.value).lower() if cell.value is not None else "" for cell in sheet[header_line] if cell.value]

    condition_col_idx = header.index(condition_col.lower()) + 1
    target_col_idx = header.index(target_col.lower()) + 1

    for row in sheet.iter_rows(min_row=header_line + 1):
        if row[condition_col_idx - 1].value == condition_val:
            row[target_col_idx - 1].value = target_val

    workbook.save(file_path)
    
@keyword('Write Cell by Column and Row')
def write_cell_by_column_and_row(file_path, sheet_name, target_col, target_row, target_val, header_line=2):
    """
    Preenche uma célula específica de uma planilha do Excel baseada em uma coluna e linha fornecidas.
    
    :param file_path: Caminho do arquivo Excel
    :param sheet_name: Nome da planilha dentro do arquivo
    :param target_col: Nome da coluna onde será preenchido o valor
    :param target_row: Número da linha onde será preenchido o valor (deve ser um número inteiro)
    :param target_val: Valor que será inserido na célula
    :param header_line: Número da linha de cabeçalho (default é 2)
    """
    # Verifica se target_row é um número inteiro
    try:
        target_row = int(target_row)
    except ValueError:
        raise TypeError(f"O valor de 'target_row' ({target_row}) deve ser um número inteiro.")

    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    try:
        sheet = workbook[sheet_name]
    except:
        raise Exception(f"sheet {sheet_name} not found")

    # Extrai os nomes das colunas a partir da linha de cabeçalho
    header = [str(cell.value).lower() if cell.value is not None else "" for cell in sheet[header_line] if cell.value]
    
    # Obtém o índice da coluna a ser preenchida
    try:
        target_col_idx = header.index(target_col.lower()) + 1  # Índice começa em 1
    except ValueError:
        raise Exception(f"Coluna '{target_col}' não encontrada na planilha")

    # Verifica se a linha é válida
    if target_row <= header_line:
        raise Exception(f"O valor de 'target_row' ({target_row}) deve ser maior que o número da linha de cabeçalho ({header_line})")

    # Preenche a célula com o valor fornecido
    sheet.cell(row=target_row, column=target_col_idx).value = target_val

    # Salva o arquivo Excel
    workbook.save(file_path)
    
@keyword('Read Excel Column')
def contar_celulas_nao_vazias(arquivo, coluna, sheetname=None):
        """
        Conta o número de células não vazias na coluna especificada do arquivo Excel, começando da linha 3.
        
        Args:
            arquivo (str): Caminho para o arquivo Excel.
            coluna (str): Coluna a ser analisada (exemplo: 'A', 'B', etc.).
            sheetname (str, opcional): Nome da planilha a ser utilizada. Se não fornecido, usa a planilha ativa.
        
        Returns:
            int: Número de células não vazias na coluna especificada.
        """
        try:
            workbook = openpyxl.load_workbook(arquivo)
            sheet = workbook[sheetname] if sheetname else workbook.active
        except Exception as e:
            raise Exception(f"Erro ao carregar o arquivo ou a planilha: {e}")

        try:
            col_idx = openpyxl.utils.column_index_from_string(coluna.upper())
        except ValueError:
            raise ValueError(f"Coluna inválida: '{coluna}'. Certifique-se de fornecer uma letra de coluna válida.")

        contagem = 0
        for row in sheet.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx, values_only=True):
            if row[0] is not None:
                contagem += 1
        return contagem

@keyword('Get Next Empty Row Number')
def get_next_empty_row_number(file_path, sheet_name, header_line=2):
    """
    Retorna o número da próxima linha vazia em uma planilha do Excel.

    :param file_path: Caminho do arquivo Excel.
    :param sheet_name: Nome da aba no Excel.
    :param header_line: Número da linha de cabeçalho (default é 2).
    :return: Número da próxima linha vazia.
    """
    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        raise Exception(f"Planilha '{sheet_name}' não encontrada no arquivo.")

    # Localiza a próxima linha vazia
    next_empty_row = header_line + 1
    while any(sheet.cell(row=next_empty_row, column=col + 1).value is not None for col in range(sheet.max_column)):
        next_empty_row += 1

    return next_empty_row

@keyword('Return Lines with data')
def listar_linhas_com_dados(arquivo, coluna, sheetname=None):
    """
    Lista as linhas que possuem dados na coluna especificada do arquivo Excel, começando da linha 3.
    
    Args:
        arquivo (str): Caminho para o arquivo Excel.
        coluna (str): Coluna a ser analisada (exemplo: 'A', 'B', etc.).
        sheetname (str, opcional): Nome da planilha a ser utilizada. Se não fornecido, usa a planilha ativa.
    
    Returns:
        list: Lista com os números das linhas que possuem dados na coluna especificada.
    """
    try:
        # Carregar o arquivo Excel e a planilha especificada
        workbook = openpyxl.load_workbook(arquivo)
        sheet = workbook[sheetname] if sheetname else workbook.active
    except Exception as e:
        raise Exception(f"Erro ao carregar o arquivo ou a planilha: {e}")

    try:
        # Obter o índice da coluna
        col_idx = column_index_from_string(coluna.upper())
    except ValueError:
        raise ValueError(f"Coluna inválida: '{coluna}'. Certifique-se de fornecer uma letra de coluna válida.")

    linhas_com_dados = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx, values_only=True), start=3):
        if row[0] is not None:
            linhas_com_dados.append(row_idx)
    
    return linhas_com_dados